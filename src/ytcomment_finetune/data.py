from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Dict

from datasets import Dataset, DatasetDict
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split


@dataclass
class LabelEncoder:
    label_to_id: Dict[str, int]
    id_to_label: Dict[int, str]


def _is_icon_char(char: str) -> bool:
    codepoint = ord(char)
    if codepoint in {0x200D, 0xFE0E, 0xFE0F}:
        return True

    emoji_blocks = (
        (0x1F1E6, 0x1F1FF),  # flags
        (0x1F300, 0x1F5FF),  # symbols and pictographs
        (0x1F600, 0x1F64F),  # emoticons
        (0x1F680, 0x1F6FF),  # transport and map symbols
        (0x1F700, 0x1F77F),
        (0x1F780, 0x1F7FF),
        (0x1F800, 0x1F8FF),
        (0x1F900, 0x1F9FF),  # supplemental symbols and pictographs
        (0x1FA70, 0x1FAFF),  # extended pictographs
        (0x2600, 0x26FF),    # misc symbols
        (0x2700, 0x27BF),    # dingbats
    )
    if any(start <= codepoint <= end for start, end in emoji_blocks):
        return True

    return unicodedata.category(char) in {"So", "Sk"}


def remove_icons(text: str) -> str:
    return "".join(char for char in text if not _is_icon_char(char))


def preprocess_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(text))
    without_urls = re.sub(r"https?://\S+|www\.\S+", " ", normalized)
    without_icons = remove_icons(without_urls)
    return re.sub(r"\s+", " ", without_icons).strip()


def load_examples(data_path: str) -> list[dict[str, str]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)

    try:
        text_index = header.index("text")
        label_index = header.index("label")
    except ValueError as exc:
        raise ValueError("The Excel file must contain 'text' and 'label' columns.") from exc

    examples: list[dict[str, str]] = []
    for row in rows:
        text = row[text_index]
        label = row[label_index]
        if text is None or label is None:
            continue

        cleaned_text = preprocess_text(text)
        cleaned_label = str(label).strip()
        if cleaned_text and cleaned_label:
            examples.append({"text": cleaned_text, "label": cleaned_label})

    if not examples:
        raise ValueError("No usable rows were found in the Excel file.")

    return examples


def build_label_encoder(examples: list[dict[str, str]]) -> LabelEncoder:
    labels = sorted({example["label"] for example in examples})
    label_to_id = {label: idx for idx, label in enumerate(labels)}
    id_to_label = {idx: label for label, idx in label_to_id.items()}
    return LabelEncoder(label_to_id=label_to_id, id_to_label=id_to_label)


def build_dataset(
    examples: list[dict[str, str]],
    label_encoder: LabelEncoder,
    test_size: float,
    seed: int,
) -> DatasetDict:
    texts = [example["text"] for example in examples]
    encoded_labels = [label_encoder.label_to_id[example["label"]] for example in examples]

    train_texts, test_texts, train_labels, test_labels = train_test_split(
        texts,
        encoded_labels,
        test_size=test_size,
        random_state=seed,
        stratify=encoded_labels,
    )

    train_dataset = Dataset.from_dict({"text": train_texts, "labels": train_labels})
    test_dataset = Dataset.from_dict({"text": test_texts, "labels": test_labels})
    return DatasetDict({"train": train_dataset, "test": test_dataset})
