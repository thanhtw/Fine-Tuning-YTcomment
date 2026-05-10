import argparse
import csv
import json
import random
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

import numpy as np
from datasets import Dataset, DatasetDict
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, classification_report, precision_recall_fscore_support
from sklearn.model_selection import train_test_split
from transformers import (
    AutoModelForSequenceClassification,
    AutoTokenizer,
    DataCollatorWithPadding,
    Trainer,
    TrainingArguments,
    set_seed,
)


DEFAULT_MODELS = [
    "bert-base-chinese",
    "hfl/chinese-roberta-wwm-ext",
]


@dataclass
class LabelEncoder:
    label_to_id: Dict[str, int]
    id_to_label: Dict[int, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fine-tune BERT and RoBERTa models on YouTube comment labels stored in an Excel file."
    )
    parser.add_argument(
        "--data-path",
        default="Data/YTcomment.xlsx",
        help="Path to the Excel file containing 'text' and 'label' columns.",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Directory for checkpoints, reports, and predictions.",
    )
    parser.add_argument(
        "--models",
        nargs="+",
        default=DEFAULT_MODELS,
        help="Hugging Face model names to fine-tune.",
    )
    parser.add_argument("--max-length", type=int, default=256, help="Maximum token length.")
    parser.add_argument("--batch-size", type=int, default=8, help="Per-device batch size.")
    parser.add_argument("--epochs", type=int, default=3, help="Number of training epochs.")
    parser.add_argument("--learning-rate", type=float, default=2e-5, help="Learning rate.")
    parser.add_argument("--weight-decay", type=float, default=0.01, help="Weight decay.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed.")
    return parser.parse_args()


def preprocess_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", str(text))
    text = re.sub(r"https?://\S+|www\.\S+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_excel_rows(data_path: str) -> List[Dict[str, str]]:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows)

    try:
        text_index = header.index("text")
        label_index = header.index("label")
    except ValueError as exc:
        raise ValueError("The Excel file must contain 'text' and 'label' columns.") from exc

    examples: List[Dict[str, str]] = []
    for row in rows:
        text = row[text_index]
        label = row[label_index]
        if text is None or label is None:
            continue

        cleaned_text = preprocess_text(text)
        cleaned_label = str(label).strip()
        if not cleaned_text or not cleaned_label:
            continue

        examples.append({"text": cleaned_text, "label": cleaned_label})

    if not examples:
        raise ValueError("No usable rows were found in the Excel file.")

    return examples


def build_label_encoder(examples: List[Dict[str, str]]) -> LabelEncoder:
    labels = sorted({example["label"] for example in examples})
    label_to_id = {label: idx for idx, label in enumerate(labels)}
    id_to_label = {idx: label for label, idx in label_to_id.items()}
    return LabelEncoder(label_to_id=label_to_id, id_to_label=id_to_label)


def build_dataset(examples: List[Dict[str, str]], label_encoder: LabelEncoder, seed: int) -> DatasetDict:
    texts = [example["text"] for example in examples]
    labels = [example["label"] for example in examples]
    encoded_labels = [label_encoder.label_to_id[label] for label in labels]

    train_texts, test_texts, train_labels, test_labels = train_test_split(
        texts,
        encoded_labels,
        test_size=0.2,
        random_state=seed,
        stratify=encoded_labels,
    )

    train_dataset = Dataset.from_dict({"text": train_texts, "labels": train_labels})
    test_dataset = Dataset.from_dict({"text": test_texts, "labels": test_labels})
    return DatasetDict({"train": train_dataset, "test": test_dataset})


def tokenize_dataset(dataset: DatasetDict, tokenizer: AutoTokenizer, max_length: int) -> DatasetDict:
    def tokenize_batch(batch: Dict[str, List[str]]) -> Dict[str, List[List[int]]]:
        return tokenizer(batch["text"], truncation=True, max_length=max_length)

    tokenized = dataset.map(tokenize_batch, batched=True, desc="Tokenizing text")
    return tokenized.remove_columns(["text"])


def compute_metrics_factory(id_to_label: Dict[int, str]):
    ordered_labels = [id_to_label[idx] for idx in sorted(id_to_label)]

    def compute_metrics(eval_pred):
        logits, labels = eval_pred
        predictions = np.argmax(logits, axis=-1)
        precision, recall, f1, _ = precision_recall_fscore_support(
            labels, predictions, average="weighted", zero_division=0
        )
        macro_precision, macro_recall, macro_f1, _ = precision_recall_fscore_support(
            labels, predictions, average="macro", zero_division=0
        )
        accuracy = accuracy_score(labels, predictions)
        report = classification_report(
            labels,
            predictions,
            labels=list(range(len(ordered_labels))),
            target_names=ordered_labels,
            zero_division=0,
            output_dict=True,
        )
        return {
            "accuracy": accuracy,
            "precision_weighted": precision,
            "recall_weighted": recall,
            "f1_weighted": f1,
            "precision_macro": macro_precision,
            "recall_macro": macro_recall,
            "f1_macro": macro_f1,
            "support_total": report["macro avg"]["support"],
        }

    return compute_metrics


def save_predictions(
    output_dir: Path,
    texts: List[str],
    true_ids: List[int],
    pred_ids: List[int],
    id_to_label: Dict[int, str],
) -> None:
    predictions_path = output_dir / "test_predictions.csv"
    with predictions_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["text", "true_label", "predicted_label"])
        writer.writeheader()
        for text, true_id, pred_id in zip(texts, true_ids, pred_ids):
            writer.writerow(
                {
                    "text": text,
                    "true_label": id_to_label[true_id],
                    "predicted_label": id_to_label[pred_id],
                }
            )


def save_report(
    output_dir: Path,
    model_name: str,
    test_texts: List[str],
    true_ids: List[int],
    pred_ids: List[int],
    id_to_label: Dict[int, str],
    metrics: Dict[str, float],
    train_count: int,
    test_count: int,
) -> None:
    ordered_labels = [id_to_label[idx] for idx in sorted(id_to_label)]
    report_text = classification_report(
        true_ids,
        pred_ids,
        labels=list(range(len(ordered_labels))),
        target_names=ordered_labels,
        zero_division=0,
    )
    report_dict = classification_report(
        true_ids,
        pred_ids,
        labels=list(range(len(ordered_labels))),
        target_names=ordered_labels,
        zero_division=0,
        output_dict=True,
    )

    metadata = {
        "model_name": model_name,
        "train_size": train_count,
        "test_size": test_count,
        "metrics": metrics,
        "labels": id_to_label,
        "classification_report": report_dict,
    }

    (output_dir / "classification_report.txt").write_text(report_text, encoding="utf-8")
    (output_dir / "metrics.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    save_predictions(output_dir, test_texts, true_ids, pred_ids, id_to_label)
    print(f"\nModel: {model_name}")
    print(f"Train/Test: {train_count}/{test_count}")
    print(report_text)


def slugify_model_name(model_name: str) -> str:
    return model_name.replace("/", "__")


def fine_tune_model(
    model_name: str,
    dataset: DatasetDict,
    label_encoder: LabelEncoder,
    args: argparse.Namespace,
) -> None:
    print(f"\nStarting fine-tuning for: {model_name}")
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    tokenized_dataset = tokenize_dataset(dataset, tokenizer, args.max_length)

    model = AutoModelForSequenceClassification.from_pretrained(
        model_name,
        num_labels=len(label_encoder.label_to_id),
        id2label=label_encoder.id_to_label,
        label2id=label_encoder.label_to_id,
    )

    model_output_dir = Path(args.output_dir) / slugify_model_name(model_name)
    model_output_dir.mkdir(parents=True, exist_ok=True)

    training_args = TrainingArguments(
        output_dir=str(model_output_dir / "checkpoints"),
        overwrite_output_dir=True,
        num_train_epochs=args.epochs,
        per_device_train_batch_size=args.batch_size,
        per_device_eval_batch_size=args.batch_size,
        learning_rate=args.learning_rate,
        weight_decay=args.weight_decay,
        evaluation_strategy="epoch",
        save_strategy="epoch",
        logging_strategy="epoch",
        load_best_model_at_end=True,
        metric_for_best_model="f1_weighted",
        greater_is_better=True,
        report_to="none",
        save_total_limit=2,
        seed=args.seed,
    )

    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=tokenized_dataset["train"],
        eval_dataset=tokenized_dataset["test"],
        tokenizer=tokenizer,
        data_collator=DataCollatorWithPadding(tokenizer=tokenizer),
        compute_metrics=compute_metrics_factory(label_encoder.id_to_label),
    )

    trainer.train()
    metrics = trainer.evaluate()
    predictions = trainer.predict(tokenized_dataset["test"])
    pred_ids = np.argmax(predictions.predictions, axis=-1).tolist()
    true_ids = predictions.label_ids.tolist()
    test_texts = dataset["test"]["text"]

    trainer.save_model(str(model_output_dir / "model"))
    tokenizer.save_pretrained(str(model_output_dir / "model"))

    serializable_metrics = {
        key: float(value) if isinstance(value, (np.floating, np.integer)) else value
        for key, value in metrics.items()
    }
    save_report(
        output_dir=model_output_dir,
        model_name=model_name,
        test_texts=test_texts,
        true_ids=true_ids,
        pred_ids=pred_ids,
        id_to_label=label_encoder.id_to_label,
        metrics=serializable_metrics,
        train_count=len(dataset["train"]),
        test_count=len(dataset["test"]),
    )


def main() -> None:
    args = parse_args()
    random.seed(args.seed)
    np.random.seed(args.seed)
    set_seed(args.seed)

    examples = load_excel_rows(args.data_path)
    label_encoder = build_label_encoder(examples)
    dataset = build_dataset(examples, label_encoder, args.seed)
    print(f"Loaded {len(examples)} usable rows from {args.data_path}")
    print(f"Train size: {len(dataset['train'])}, Test size: {len(dataset['test'])}")
    print(f"Labels: {label_encoder.label_to_id}")

    output_root = Path(args.output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "label_mapping.json").write_text(
        json.dumps(label_encoder.label_to_id, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    for model_name in args.models:
        fine_tune_model(model_name, dataset, label_encoder, args)


if __name__ == "__main__":
    main()
